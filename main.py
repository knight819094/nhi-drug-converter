"""
健保用藥品項轉換工具
執行方式：
  python main.py          → 開啟圖形介面
  python main.py --cli    → 命令列模式（無 GUI）
"""

import re
import subprocess
import sys
import threading
import tkinter as tk
import webbrowser
from datetime import date
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

# Windows 上 "Helvetica" → "Arial"，其他平台維持
_FONT = "Arial" if sys.platform == "win32" else "Helvetica"

def _load_font_size() -> int:
    try:
        with open(Path(__file__).parent / "config.yaml", encoding="utf-8") as f:
            cfg = yaml.safe_load(f)
        return int(cfg.get("ui", {}).get("font_size", 10))
    except Exception:
        return 10

_FS   = _load_font_size()   # 基礎字級
_FSs  = max(_FS - 1, 7)     # 小字（狀態列、標籤）
_FSl  = _FS + 1             # 大字（區塊標題）
_FSxl = _FS + 3             # 特大（按鈕）
_FSh  = _FS + 5             # 標題列

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import yaml

# ════════════════════════════════════════════════════════════
#  設定
# ════════════════════════════════════════════════════════════
BASE_DIR  = Path(__file__).parent
_cfg_cache: dict | None = None

def get_config() -> dict:
    global _cfg_cache
    if _cfg_cache is None:
        with open(BASE_DIR / "config.yaml", encoding="utf-8") as f:
            _cfg_cache = yaml.safe_load(f)
    return _cfg_cache

def cfg_path(key: str) -> Path:
    return BASE_DIR / get_config()["paths"][key]

# ════════════════════════════════════════════════════════════
#  主題色彩
# ════════════════════════════════════════════════════════════
THEMES = {
    "light": {
        "bg":           "#F5F7FA",
        "panel":        "#FFFFFF",
        "accent":       "#1F4E79",
        "accent_lt":    "#2E75B6",
        "text":         "#1A1A1A",
        "gray":         "#6B7280",
        "border":       "#D1D5DB",
        "link_bar":     "#E8EEF4",
        "log_bg":       "#1E1E1E",
        "log_fg":       "#D4D4D4",
        "btn_run":      "#217346",
        "btn_sm":       "#4B5563",
        "tab_act":      "#FFFFFF",
        "tab_inact":    "#DDE6EF",
        "match_bg":     "#FFF9C4",
        "entry_bg":     "#F9FAFB",
        "drop_border":  "#2E75B6",
    },
    "dark": {
        "bg":           "#1A1A2E",
        "panel":        "#2D2D44",
        "accent":       "#5BA4E8",
        "accent_lt":    "#7DBFFF",
        "text":         "#E0E0F0",
        "gray":         "#9CA3AF",
        "border":       "#4B5563",
        "link_bar":     "#252540",
        "log_bg":       "#0D0D1A",
        "log_fg":       "#A8E6CF",
        "btn_run":      "#1A7A45",
        "btn_sm":       "#374151",
        "tab_act":      "#2D2D44",
        "tab_inact":    "#1A1A2E",
        "match_bg":     "#4A4A00",
        "entry_bg":     "#252540",
        "drop_border":  "#5BA4E8",
    },
}
def _detect_system_theme() -> str:
    """偵測作業系統目前的深色/淺色模式。"""
    try:
        if sys.platform == "darwin":
            result = subprocess.run(
                ["defaults", "read", "-g", "AppleInterfaceStyle"],
                capture_output=True, text=True
            )
            return "dark" if result.stdout.strip() == "Dark" else "light"
        elif sys.platform == "win32":
            import winreg
            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize"
            )
            value, _ = winreg.QueryValueEx(key, "AppsUseLightTheme")
            winreg.CloseKey(key)
            return "dark" if value == 0 else "light"
    except Exception:
        pass
    return "light"

_theme_name = _detect_system_theme()

def T(key: str) -> str:
    return THEMES[_theme_name][key]

# ════════════════════════════════════════════════════════════
#  解析常數
# ════════════════════════════════════════════════════════════
CONC_PAT = re.compile(
    r"(\d[\d,]*\.?\d*\s*"
    r"(?:MG/ML|MCG/ML|MG/G|G/ML|MG|MCG|USP-U|FIP-U|BP-U|USP"
    r"|G|IU/ML|IU/G|IU|UNIT|%)[^\s]*)"
)
MANUF_PAT = re.compile(
    r"[一-鿿]{2,}(?:股份有限公司|有限公司|製藥廠|股份有限|藥品|藥廠|有限|公司|廠)[一-鿿]*"
)
# 中文藥名前置廠商名（後接空格）
_MANUF_PREFIX_PAT = re.compile(
    r"^([一-鿿]{2,}(?:股份有限公司|有限公司|製藥廠|股份有限|藥品|藥廠|有限|公司|廠)[一-鿿]*)\s+"
)
# 前置引號廠商名，如 "鴻汶"安保喜樂...
_QUOTED_PREFIX_PAT = re.compile(r'^"([^"]+)"(.+)')
_FW2HW = str.maketrans(
    "〝〞！＂＃＄％＆＇（）＊＋，－．／０１２３４５６７８９：；＜＝＞？＠"
    "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ［＼］＾＿｀"
    "ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ｛｜｝～　",
    "\"\"!\"#$%&'()*+,-./0123456789:;<=>?@"
    "ABCDEFGHIJKLMNOPQRSTUVWXYZ[\\]^_`"
    "abcdefghijklmnopqrstuvwxyz{|}~ "
)
ROC_OFFSET = 1911
TODAY      = date.today()

OUTPUT_COLS = [
    "藥品代碼", "中文藥名", "建議", "健保價格", "生效日", "停用日",
    "英文藥名", "主成分", "劑型", "廠商", "含量", "容量",
    "英文藥名首位", "主成分首位",
]

# ════════════════════════════════════════════════════════════
#  核心函式
# ════════════════════════════════════════════════════════════
def to_halfwidth(s: str) -> str:
    return s.translate(_FW2HW) if s else s

def strip_num_zeros(s: str) -> str:
    """移除數值中多餘的尾隨零：5.000 MG/ML → 5 MG/ML，12.50 MG → 12.5 MG"""
    return re.sub(r'\d+\.\d+', lambda m: m.group(0).rstrip('0').rstrip('.'), s)

def clean_cn_name(name: str) -> str:
    """全形轉半形後，將前置廠商名移至末尾。"""
    name = to_halfwidth(name)
    # 優先處理引號廠商名：如 "鴻汶"安保喜樂... → 安保喜樂..."鴻汶"
    m = _QUOTED_PREFIX_PAT.match(name)
    if m:
        rest = m.group(2).strip()
        if rest:
            return f'{rest}"{m.group(1)}"'
    # 裸字廠商名：如 台裕製藥廠 某某錠 → 某某錠 台裕製藥廠
    m = _MANUF_PREFIX_PAT.match(name)
    if m:
        rest = name[m.end():].strip()
        if rest:
            return f"{rest} {m.group(1)}"
    return name

def roc_to_date(roc_str: str) -> date | None:
    s = roc_str.strip()
    if len(s) < 7:
        return None
    try:
        return date(int(s[:3]) + ROC_OFFSET, int(s[3:5]), int(s[5:7]))
    except (ValueError, TypeError):
        return None

def _manufacturer(line: str) -> str:
    if len(line) > 1700:
        m = MANUF_PAT.search(" ".join(line[1700:].split()))
        if m:
            return m.group()
    m = MANUF_PAT.search(" ".join(line[440:720].split()))
    return m.group() if m else line[600:660].strip()

def parse_files(txt_files: list[Path], log_fn=print,
                progress_fn=None, history_mode=False) -> pd.DataFrame:
    """
    history_mode=False：只保留停用日 >= 今天的紀錄，每個藥品代碼取最新生效那筆。
    history_mode=True ：保留所有價格異動歷史，不去重。
    progress_fn(done_lines, total_lines) 每 500 行回呼一次。
    """
    all_lines: list[tuple[Path, list[str]]] = []
    for fp in txt_files:
        text = fp.read_bytes().decode("big5hkscs", errors="replace")
        all_lines.append((fp, text.splitlines()))
    total_lines = sum(len(lines) for _, lines in all_lines)
    done_lines  = 0

    frames = []
    for fp, splitlines in all_lines:
        log_fn(f"  讀取：{fp.name}")
        records = []
        for i, line in enumerate(splitlines):
            if len(line) < 54:
                continue
            end_raw, price_raw = line[46:54].strip(), line[27:37].strip()
            if not history_mode:
                # 非歷史模式：排除已過期紀錄
                end_dt = roc_to_date(end_raw)
                if end_dt is not None and end_dt < TODAY:
                    continue
            try:
                price = float(price_raw)
            except ValueError:
                continue
            if price < 0:
                continue

            cn_m      = re.search(r"\d\s{2,}(.+)", line[750:870])
            cn_name   = clean_cn_name(
                cn_m.group(1).strip() if cn_m else line[750:870].strip()
            )
            seg       = " ".join(line[200:600].split())
            conc_m    = CONC_PAT.search(seg)
            ingredient = seg[:conc_m.start()].strip() if conc_m else seg.strip()
            conc       = conc_m.group().strip()        if conc_m else ""
            dosage_m  = re.search(r"[一-鿿！-￮]+", " ".join(line[320:460].split()))
            dosage    = dosage_m.group() if dosage_m else ""
            atc = ""
            if len(line) > 1700:
                atc_m = re.search(r"([A-Z]\d{2}[A-Z]{2}\d{2})\s*$", line[1700:].strip())
                if atc_m:
                    atc = atc_m.group(1)

            hw = to_halfwidth
            records.append({
                "藥品代碼":   line[17:27].strip(),
                "中文藥名":   cn_name,
                "英文藥名":   hw(line[54:154].strip()),
                "劑型":       hw(dosage),
                "主成分":     hw(ingredient),
                "含量":       strip_num_zeros(hw(conc)),
                "容量":       strip_num_zeros(hw(line[154:200].strip())),
                "廠商":       hw(_manufacturer(line)).removeprefix("台灣"),
                "ATC碼":     atc,
                "健保價格":   price,
                "生效日_raw": line[38:46].strip(),
                "停用日_raw": line[46:54].strip(),
            })
            done_lines += 1
            if progress_fn and done_lines % 500 == 0:
                progress_fn(done_lines, total_lines)
        if progress_fn:
            progress_fn(done_lines, total_lines)
        frames.append(pd.DataFrame(records))

    df = pd.concat(frames, ignore_index=True)
    df = df.sort_values("生效日_raw")
    if not history_mode:
        # 每個藥品代碼只保留最新生效的那筆
        df = df.drop_duplicates(subset=["藥品代碼"], keep="last")
    df["生效日"] = df["生效日_raw"].apply(roc_to_date)
    df["停用日"] = df["停用日_raw"].apply(roc_to_date)
    df = df.drop(columns=["生效日_raw", "停用日_raw"])
    return df.reset_index(drop=True)

def add_recommendation(df: pd.DataFrame) -> pd.DataFrame:
    new_days = get_config()["recommendation_rules"]["new_item_days"]
    def label(row):
        start  = row["生效日"]
        is_new = (start and isinstance(start, date)
                  and 0 <= (TODAY - start).days <= new_days)
        if is_new:        return "🆕 本月新增"
        if row["健保價格"] == 0: return "❌ 無健保給付"
        return "✅ 正常"
    def _first_word(s):
        s = str(s).strip() if pd.notna(s) else ""
        return s.split()[0] if s else ""

    df = df.copy()
    df["建議"]     = df.apply(label, axis=1)
    df["英文藥名首位"] = df["英文藥名"].apply(_first_word)
    df["主成分首位"]   = df["主成分"].apply(lambda s: _first_word(s).title())
    return df

def month_str_from_files(txt_files: list[Path]) -> str:
    m = re.search(r"_(\d{5})_", txt_files[0].name) if txt_files else None
    return f"{int(m.group(1)[:3]) + 1911}{m.group(1)[3:]}" if m else "未知月份"

# ── Excel 匯出 ───────────────────────────────────────────────
_SUGGEST_FILL = {
    "🆕 本月新增":   PatternFill("solid", fgColor="B3FFB3"),
    "❌ 無健保給付": PatternFill("solid", fgColor="FFD9B3"),
    "✅ 正常":       None,
}
_HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT  = Font(color="FFFFFF", bold=True, size=11)
_NORM_FONT = Font(size=10)
_CENTER    = Alignment(horizontal="center", vertical="center")
_RIGHT     = Alignment(horizontal="right",  vertical="center")
_LEFT      = Alignment(horizontal="left",   vertical="center")
_PRICE_IDX = OUTPUT_COLS.index("健保價格") + 1
_SUGG_IDX  = len(OUTPUT_COLS)

def export_excel(df: pd.DataFrame, month_str: str,
                 out_dir: Path | None = None,
                 progress_fn=None) -> tuple[Path, int, int]:
    cfg      = get_config()
    out_dir  = out_dir or cfg_path("output_dir")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"健保用藥_{month_str}.xlsx"

    alt_fill   = PatternFill("solid", fgColor=cfg["output"]["alternating_row_color"])
    max_width  = cfg["output"]["max_col_width"]
    sugg_width = cfg["output"]["suggest_col_width"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"全品項_{month_str}"

    ws.append(OUTPUT_COLS)
    for ci, _ in enumerate(OUTPUT_COLS, 1):
        c = ws.cell(1, ci)
        c.fill, c.font, c.alignment = _HDR_FILL, _HDR_FONT, _CENTER

    total_rows = len(df)
    for ri, row in enumerate(df[OUTPUT_COLS].itertuples(index=False), 2):
        alt, suggest = ri % 2 == 0, row[-1]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            if isinstance(val, date):
                cell.value, cell.alignment = val.strftime("%Y/%m/%d"), _CENTER
            elif ci == _PRICE_IDX:
                cell.value, cell.number_format, cell.alignment = val, "#,##0.00", _RIGHT
            else:
                cell.value, cell.alignment = val, _LEFT
            cell.font = _NORM_FONT
            if ci == _SUGG_IDX:
                fill = _SUGGEST_FILL.get(suggest)
                cell.fill = fill if fill else (alt_fill if alt else PatternFill())
            elif alt:
                cell.fill = alt_fill
        if progress_fn and ri % 200 == 0:
            progress_fn(ri - 1, total_rows)

    for ci, col_name in enumerate(OUTPUT_COLS, 1):
        letter  = get_column_letter(ci)
        max_len = max((len(str(c.value or "")) for c in ws[letter][1:]),
                      default=len(col_name))
        ws.column_dimensions[letter].width = (
            sugg_width if col_name == "建議" else min(max_len + 2, max_width)
        )

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2   = wb.create_sheet("摘要")
    total = len(df)
    new_c = int((df["建議"] == "🆕 本月新增").sum())
    for row in [
        ["項目", "數量"],
        ["總品項數",   total],
        ["本月新增",   new_c],
        ["正常品項",   int((df["建議"] == "✅ 正常").sum())],
        ["無健保給付", int((df["建議"] == "❌ 無健保給付").sum())],
        ["轉換日期",   TODAY.strftime("%Y/%m/%d")],
    ]:
        ws2.append(row)
    for cell in ws2[1]:
        cell.fill, cell.font = _HDR_FILL, _HDR_FONT

    wb.save(out_path)
    return out_path, total, new_c

# ════════════════════════════════════════════════════════════
#  GUI 元件
# ════════════════════════════════════════════════════════════

def open_path(p: Path):
    """跨平台開啟檔案或資料夾。"""
    if sys.platform == "win32":
        import os
        os.startfile(str(p))
    elif sys.platform == "darwin":
        subprocess.call(["open", str(p)])
    else:
        subprocess.call(["xdg-open", str(p)])


def make_btn(parent, text, command, bg, fg="white",
             font_size=None, bold=False, padx=12, pady=5) -> tk.Label:
    """macOS-compatible 彩色按鈕（tk.Label + 點擊事件）。"""
    weight = "bold" if bold else "normal"
    fs = font_size if font_size is not None else _FS
    btn = tk.Label(
        parent, text=text,
        bg=bg, fg=fg,
        font=(_FONT, fs, weight),
        padx=padx, pady=pady,
        cursor="hand2", relief="flat",
    )
    def _press(e):
        btn.configure(bg=_darken(bg))
    def _release(e):
        btn.configure(bg=bg)
        command()
    btn.bind("<ButtonPress-1>",   _press)
    btn.bind("<ButtonRelease-1>", _release)
    return btn

def _darken(hex_color: str, factor=0.82) -> str:
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return "#{:02X}{:02X}{:02X}".format(
        int(r * factor), int(g * factor), int(b * factor))


class SuccessModal(tk.Toplevel):
    """轉換完成 Modal。"""
    def __init__(self, parent, total: int, new_c: int, out_path: Path):
        super().__init__(parent)
        self.title("轉換完成")
        self.resizable(False, False)
        self.configure(bg=T("panel"))
        self.grab_set()

        # 置中於主視窗
        self.update_idletasks()
        px, py = parent.winfo_rootx(), parent.winfo_rooty()
        pw, ph = parent.winfo_width(), parent.winfo_height()
        self.geometry(f"360x220+{px + pw//2 - 180}+{py + ph//2 - 110}")

        tk.Label(self, text="✅", font=(_FONT, _FS * 36 // 10),
                 bg=T("panel"), fg="#217346").pack(pady=(24, 4))
        tk.Label(self, text="轉換成功！",
                 font=(_FONT, _FSh - 1, "bold"),
                 bg=T("panel"), fg=T("text")).pack()
        tk.Label(self,
                 text=f"共處理 {total:,} 筆資料　本月新增 {new_c} 筆",
                 font=(_FONT, _FS),
                 bg=T("panel"), fg=T("gray")).pack(pady=(4, 16))

        btn_row = tk.Frame(self, bg=T("panel"))
        btn_row.pack()

        def open_excel():
            open_path(out_path)
            self.destroy()

        make_btn(btn_row, "📂 開啟 Excel", open_excel,
                 bg="#217346", font_size=_FSl, bold=True,
                 padx=18, pady=8).pack(side="left", padx=8)
        make_btn(btn_row, "  關閉  ", self.destroy,
                 bg="#4B5563", font_size=_FSl,
                 padx=18, pady=8).pack(side="left", padx=8)


class CustomProgressBar(tk.Canvas):
    """依百分比顯示的進度條，完成後變綠。"""
    BAR_H = 18

    def __init__(self, parent, **kw):
        super().__init__(parent, height=self.BAR_H,
                         bg=T("bg"), highlightthickness=0, **kw)
        self._anim_id = None
        self._running = False

    def _bw(self): return self.winfo_width() or 680
    def _bh(self): return self.winfo_height() or self.BAR_H

    def set_value(self, done: int, total: int):
        """依 done/total 繪製進度，線程安全（需在主執行緒呼叫）。"""
        if self._running:
            return
        if self._anim_id:
            self.after_cancel(self._anim_id)
            self._anim_id = None
        pct = min(done / total, 1.0) if total > 0 else 0
        w, h = self._bw(), self._bh()
        self.delete("all")
        self.create_rectangle(0, 0, w, h, fill=T("border"), outline="", tags="bg")
        fill_w = int(w * pct)
        if fill_w > 0:
            self.create_rectangle(0, 0, fill_w, h,
                                  fill=T("accent_lt"), outline="", tags="bar")
        self.create_text(w // 2, h // 2, text=f"{int(pct * 100)}%",
                         fill=T("text"), font=(_FONT, _FSs, "bold"), tags="pct")

    def stop_success(self):
        if self._anim_id:
            self.after_cancel(self._anim_id)
            self._anim_id = None
        self._running = False
        w, h = self._bw(), self._bh()
        self.delete("all")
        self.create_rectangle(0, 0, w, h, fill="#217346", outline="")
        self.create_text(w // 2, h // 2, text="100%  完成！",
                         fill="white", font=(_FONT, _FSs, "bold"))

    def reset(self):
        self._running = False
        if self._anim_id:
            self.after_cancel(self._anim_id)
            self._anim_id = None
        self.delete("all")
        w, h = self._bw(), self._bh()
        self.create_rectangle(0, 0, w, h, fill=T("border"), outline="")


class ScrollableCheck(tk.Frame):
    """搜尋框（含符合結果高亮）+ 捲動三欄 Checkbox 清單。"""
    COLS = 3

    def __init__(self, parent, height=200, **kw):
        super().__init__(parent, bg=T("panel"), **kw)

        # 搜尋框
        sr = tk.Frame(self, bg=T("panel"))
        sr.pack(fill="x", padx=6, pady=(6, 2))
        tk.Label(sr, text="🔍", bg=T("panel"),
                 font=(_FONT, _FSl + 1)).pack(side="left")
        self._sv = tk.StringVar()
        self._sv.trace_add("write", lambda *_: self._filter())
        self._search_entry = tk.Entry(
            sr, textvariable=self._sv, font=(_FONT, _FSl),
            relief="solid", bd=1, fg=T("text"), bg=T("entry_bg")
        )
        self._search_entry.pack(side="left", fill="x", expand=True, padx=(4, 0))

        # 結果數
        self._result_lbl = tk.Label(sr, text="", bg=T("panel"),
                                    fg=T("gray"), font=(_FONT, _FSs))
        self._result_lbl.pack(side="left", padx=(6, 0))

        canvas = tk.Canvas(self, bg=T("panel"), highlightthickness=0, height=height)
        sb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self._inner  = tk.Frame(canvas, bg=T("panel"))
        self._win    = canvas.create_window((0, 0), window=self._inner, anchor="nw")
        self._canvas = canvas

        self._inner.bind("<Configure>",
                         lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfig(self._win, width=e.width))
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))

        self._all:  list[str]             = []
        self._vars: dict[str, tk.BooleanVar] = {}

    def populate(self, items: list[str], default=True):
        self._all = items
        for lbl in items:
            self._vars.setdefault(lbl, tk.BooleanVar(value=default))
        self._render(items)

    def _filter(self):
        kw = self._sv.get().strip()
        matched = [l for l in self._all if kw in l] if kw else self._all
        self._render(matched, kw)

    def _render(self, items: list[str], kw: str = ""):
        for w in self._inner.winfo_children():
            w.destroy()
        for i, lbl in enumerate(items):
            var     = self._vars.setdefault(lbl, tk.BooleanVar(value=True))
            is_match = bool(kw and kw in lbl)
            cb = tk.Checkbutton(
                self._inner, text=lbl, variable=var,
                bg=T("match_bg") if is_match else T("panel"),
                activebackground=T("match_bg") if is_match else T("panel"),
                fg=T("accent") if is_match else T("text"),
                font=(_FONT, _FS, "bold") if is_match else (_FONT, _FS),
                anchor="w", relief="flat", bd=0, cursor="hand2",
            )
            cb.grid(row=i // self.COLS, column=i % self.COLS,
                    sticky="w", padx=6, pady=1)
        self._canvas.yview_moveto(0)
        # 更新結果數
        if kw:
            self._result_lbl.configure(text=f"{len(items)} 項")
        else:
            self._result_lbl.configure(text="")

    def select_all(self, val=True):
        kw = self._sv.get().strip()
        for lbl in (self._all if not kw else [l for l in self._all if kw in l]):
            self._vars[lbl].set(val)
        self._filter()

    def selected(self) -> list[str]:
        return [k for k, v in self._vars.items() if v.get()]


class DropZoneEntry(tk.Frame):
    """帶拖曳視覺提示的路徑輸入框。"""
    def __init__(self, parent, textvariable, width=48, **kw):
        super().__init__(parent, bg=T("panel"), **kw)
        self._var = textvariable
        self._focused = False

        self._entry = tk.Entry(
            self, textvariable=textvariable, width=width,
            font=(_FONT, _FS), relief="flat", bd=0,
            fg=T("text"), bg=T("entry_bg"),
        )
        self._entry.pack(fill="x", padx=2, pady=2)

        self.configure(
            highlightbackground=T("border"),
            highlightcolor=T("drop_border"),
            highlightthickness=1,
        )
        self._entry.bind("<FocusIn>",  self._on_focus)
        self._entry.bind("<FocusOut>", self._on_blur)

        # 拖曳提示標籤（顯示在右側）
        self._hint = tk.Label(
            self._entry, text="  拖曳資料夾至此",
            font=(_FONT, _FSs), fg=T("gray"),
            bg=T("entry_bg"), cursor="arrow",
        )
        # 只有空白時才顯示
        textvariable.trace_add("write", self._update_hint)

    def _on_focus(self, _=None):
        self.configure(highlightbackground=T("drop_border"),
                       highlightthickness=2)

    def _on_blur(self, _=None):
        self.configure(highlightbackground=T("border"),
                       highlightthickness=1)

    def _update_hint(self, *_):
        if self._var.get():
            self._hint.place_forget()
        else:
            self._hint.place(relx=1.0, rely=0.5, anchor="e", x=-4)


# ════════════════════════════════════════════════════════════
#  主視窗
# ════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("健保用藥品項轉換工具")
        self.geometry("720x740")
        self.resizable(False, False)
        self.configure(bg=T("bg"))

        self.input_dir  = tk.StringVar(value=str(cfg_path("input_dir")))
        self.output_dir = tk.StringVar(value=str(cfg_path("output_dir")))
        self.running    = False
        self._df_cache  = None
        self._dosage_map: dict[str, str] = {}
        self._last_out_path: Path | None = None

        self._build()
        self._scan_async()

    # ── 建構 ────────────────────────────────────────────────
    def _build(self):
        self._build_header()
        self._build_link_bar()
        self._build_tabs()

    def _build_header(self):
        hdr = tk.Frame(self, bg=T("accent"), height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        tk.Label(hdr, text="  健保用藥品項轉換工具",
                 bg=T("accent"), fg="white",
                 font=(_FONT, _FSh, "bold")).pack(side="left", padx=6, pady=10)

        # 右側：設定齒輪
        right = tk.Frame(hdr, bg=T("accent"))
        right.pack(side="right", padx=10)

        gear = tk.Label(right, text="⚙️", bg=T("accent"),
                        font=(_FONT, _FSh + 1), cursor="hand2")
        gear.pack(side="left")
        gear.bind("<Button-1>", lambda _: self._open_config())

    def _build_link_bar(self):
        self._link_bar = tk.Frame(self, bg=T("link_bar"))
        self._link_bar.pack(fill="x")
        for text, url in [
            ("🔍 健保用藥查詢", "https://info.nhi.gov.tw/INAE3000/INAE3000S01"),
            ("⬇️  下載更新檔",   "https://www.nhi.gov.tw/ch/lp-2466-1.html"),
        ]:
            lbl = tk.Label(self._link_bar, text=text,
                           bg=T("link_bar"), fg=T("accent"),
                           font=(_FONT, _FS, "underline"),
                           cursor="hand2", padx=14, pady=5)
            lbl.pack(side="left")
            lbl.bind("<Button-1>", lambda e, u=url: webbrowser.open(u))
            lbl.bind("<Enter>",    lambda e, b=lbl: b.configure(fg=T("accent_lt")))
            lbl.bind("<Leave>",    lambda e, b=lbl: b.configure(fg=T("accent")))

    def _build_tabs(self):
        # Tab 列
        tab_bar = tk.Frame(self, bg=T("bg"))
        tab_bar.pack(fill="x", padx=0)
        self._tab_frames: dict[str, tk.Frame] = {}
        self._tab_btns:   dict[str, tk.Label] = {}

        self._content = tk.Frame(self, bg=T("bg"))
        self._content.pack(fill="both", expand=True)

        for tab_name in ["資料轉換", "歷史紀錄"]:
            btn = tk.Label(tab_bar, text=f"  {tab_name}  ",
                           bg=T("tab_inact"), fg=T("gray"),
                           font=(_FONT, _FSl), cursor="hand2",
                           padx=4, pady=6, relief="flat")
            btn.pack(side="left")
            btn.bind("<Button-1>", lambda e, n=tab_name: self._switch_tab(n))
            self._tab_btns[tab_name] = btn

            frame = tk.Frame(self._content, bg=T("bg"))
            self._tab_frames[tab_name] = frame

        self._build_convert_tab(self._tab_frames["資料轉換"])
        self._build_history_tab(self._tab_frames["歷史紀錄"])
        self._switch_tab("資料轉換")

    def _switch_tab(self, name: str):
        for n, f in self._tab_frames.items():
            f.pack_forget()
        for n, b in self._tab_btns.items():
            active = (n == name)
            b.configure(
                bg=T("tab_act") if active else T("tab_inact"),
                fg=T("accent") if active else T("gray"),
                font=(_FONT, _FSl, "bold") if active else (_FONT, _FSl),
            )
        self._tab_frames[name].pack(fill="both", expand=True)
        if name == "歷史紀錄":
            self._refresh_history()

    # ── 資料轉換 Tab ────────────────────────────────────────
    def _build_convert_tab(self, parent):
        body = tk.Frame(parent, bg=T("bg"))
        body.pack(fill="both", expand=True, padx=20, pady=12)

        # 路徑卡片
        self._section(body, "📁 資料來源與輸出位置")
        pc = self._card(body)
        self._path_row(pc, "TXT 來源資料夾",   self.input_dir,  self._browse_in,  0)
        self._path_row(pc, "Excel 輸出資料夾",  self.output_dir, self._browse_out, 1)

        # 快速篩選
        self._section(body, "⚡ 快速篩選")
        qf = self._card(body)

        # 第一行：三選一（互斥）
        self._history_mode  = tk.BooleanVar(value=False)
        self._filter_new    = tk.BooleanVar(value=False)
        self._filter_future = tk.BooleanVar(value=False)
        # 第二行：獨立複選
        self._filter_active = tk.BooleanVar(value=True)   # 只顯示停用日 > 今日
        self._filter_nopay  = tk.BooleanVar(value=False)

        _row1_vars = [self._history_mode, self._filter_new, self._filter_future]
        self._filter_lock = False

        def _make_exclusive(selected_var):
            def _cb(*_):
                if self._filter_lock:
                    return
                if selected_var.get():
                    self._filter_lock = True
                    for v in _row1_vars:
                        if v is not selected_var:
                            v.set(False)
                    self._filter_lock = False
            return _cb

        for v in _row1_vars:
            v.trace_add("write", _make_exclusive(v))

        def _ck(parent, text, variable, padx=(0, 22)):
            tk.Checkbutton(
                parent, text=text, variable=variable,
                bg=T("panel"), fg=T("text"),
                activebackground=T("panel"),
                selectcolor=T("panel"),
                font=(_FONT, _FS),
                cursor="hand2", relief="flat", bd=0,
            ).pack(side="left", padx=padx)

        # ── 第一行 ──
        row1 = tk.Frame(qf, bg=T("panel"))
        row1.pack(anchor="w", pady=(0, 6))
        _ck(row1, "📋 輸出全部歷史資料",   self._history_mode)
        _ck(row1, "🆕 只輸出當月更新",     self._filter_new)
        _ck(row1, "🔮 匯出半年內生效資料", self._filter_future, padx=(0, 0))

        # ── 第二行 ──
        row2 = tk.Frame(qf, bg=T("panel"))
        row2.pack(anchor="w")
        _ck(row2, "📅 只顯示停用日大於今日", self._filter_active)
        _ck(row2, "❌ 排除無健保給付",        self._filter_nopay, padx=(0, 0))

        # 劑型篩選
        self._section(body, "💊 劑型篩選（勾選要保留的劑型）")
        br = tk.Frame(body, bg=T("bg"))
        br.pack(anchor="w", pady=(0, 4))
        for txt, val in [("全選", True), ("全不選", False)]:
            make_btn(br, txt,
                     command=lambda v=val: self.checklist.select_all(v),
                     bg="#2E75B6", font_size=_FSs,
                     padx=10, pady=3).pack(side="left", padx=(0, 6))

        lf = tk.Frame(body, bg=T("panel"), bd=1, relief="solid",
                      highlightbackground=T("border"), highlightthickness=1)
        lf.pack(fill="x")
        self.checklist = ScrollableCheck(lf, height=200)
        self.checklist.pack(fill="both", padx=4, pady=4)

        self.status_lbl = tk.Label(body, text="正在掃描劑型...",
                                   bg=T("bg"), fg=T("gray"),
                                   font=(_FONT, _FSs))
        self.status_lbl.pack(anchor="w", pady=(2, 0))

        # 執行按鈕
        ba = tk.Frame(parent, bg=T("bg"))
        ba.pack(pady=(4, 6))
        self.run_btn = make_btn(
            ba, "▶   開始轉換", self._run,
            bg="#D4A017", fg="#1A1A1A", font_size=_FSxl, bold=True,
            padx=40, pady=12,
        )
        self.run_btn.pack()

        # 進度條
        self.prog_bar = CustomProgressBar(parent)
        self.prog_bar.pack(fill="x", padx=20, pady=(0, 4))
        self.after(100, self.prog_bar.reset)

        # 日誌
        lw = tk.Frame(parent, bg=T("bg"))
        lw.pack(fill="both", expand=True, padx=20, pady=(0, 14))
        self.log = tk.Text(lw, height=6,
                           bg=T("log_bg"), fg=T("log_fg"),
                           font=("Courier", 10), relief="flat", bd=0,
                           state="disabled")
        self.log.pack(fill="both", expand=True)

    # ── 歷史紀錄 Tab ────────────────────────────────────────
    def _build_history_tab(self, parent):
        self._section_in(parent, "📋 已匯出的 Excel 紀錄")
        self._hist_frame = tk.Frame(parent, bg=T("bg"))
        self._hist_frame.pack(fill="both", expand=True, padx=20, pady=4)

    def _refresh_history(self):
        for w in self._hist_frame.winfo_children():
            w.destroy()

        out_dir = Path(self.output_dir.get())
        files   = sorted(out_dir.glob("健保用藥_*.xlsx"), reverse=True)

        if not files:
            tk.Label(self._hist_frame,
                     text="尚無匯出紀錄",
                     bg=T("bg"), fg=T("gray"),
                     font=(_FONT, _FSl)).pack(pady=40)
            return

        for fp in files:
            row = tk.Frame(self._hist_frame, bg=T("panel"),
                           pady=6, padx=10,
                           highlightbackground=T("border"),
                           highlightthickness=1)
            row.pack(fill="x", pady=3)

            size_kb = fp.stat().st_size // 1024
            mod     = date.fromtimestamp(fp.stat().st_mtime).strftime("%Y/%m/%d")

            tk.Label(row, text=fp.name,
                     bg=T("panel"), fg=T("text"),
                     font=(_FONT, _FS, "bold")).pack(side="left")
            tk.Label(row, text=f"  {size_kb} KB　{mod}",
                     bg=T("panel"), fg=T("gray"),
                     font=(_FONT, _FSs)).pack(side="left")

            def open_file(p=fp):
                open_path(p)

            make_btn(row, "  開啟  ", open_file,
                     bg="#2E75B6", font_size=_FSs,
                     padx=10, pady=4).pack(side="right", padx=4)

    # ── 輔助元件 ────────────────────────────────────────────
    def _section(self, parent, text):
        tk.Label(parent, text=text, bg=T("bg"), fg=T("accent"),
                 font=(_FONT, _FSl, "bold")).pack(anchor="w", pady=(8, 2))

    def _section_in(self, parent, text):
        tk.Label(parent, text=text, bg=T("bg"), fg=T("accent"),
                 font=(_FONT, _FSl, "bold")).pack(anchor="w",
                                                      padx=20, pady=(12, 4))

    def _card(self, parent):
        f = tk.Frame(parent, bg=T("panel"), bd=1, relief="solid",
                     highlightbackground=T("border"), highlightthickness=1)
        f.pack(fill="x", pady=(0, 8))
        inner = tk.Frame(f, bg=T("panel"))
        inner.pack(fill="x", padx=12, pady=8)
        return inner

    def _path_row(self, parent, label, var, cmd, row):
        tk.Label(parent, text=label, bg=T("panel"), fg=T("gray"),
                 font=(_FONT, _FS)).grid(row=row, column=0, sticky="w", pady=4)
        entry = DropZoneEntry(parent, textvariable=var, width=44)
        entry.grid(row=row, column=1, padx=(8, 4), pady=2)
        make_btn(parent, "  瀏覽  ", cmd,
                 bg="#2E75B6", font_size=_FSs,
                 padx=8, pady=4).grid(row=row, column=2, padx=(4, 0))

    def _log(self, msg):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    # ── 操作 ────────────────────────────────────────────────
    def _open_config(self):
        open_path(BASE_DIR / "config.yaml")

    def _browse_in(self):
        d = filedialog.askdirectory(initialdir=self.input_dir.get())
        if d:
            self.input_dir.set(d)
            self._df_cache = None
            self.checklist.populate([], True)
            self._scan_async()

    def _browse_out(self):
        d = filedialog.askdirectory(initialdir=self.output_dir.get())
        if d:
            self.output_dir.set(d)

    # ── 掃描劑型 ────────────────────────────────────────────
    def _scan_async(self):
        threading.Thread(target=self._scan_worker, daemon=True).start()

    def _scan_worker(self):
        try:
            p     = Path(self.input_dir.get())
            files = sorted(p.glob("*.TXT")) + sorted(p.glob("*.txt"))
            if not files:
                self.after(0, lambda: self.status_lbl.configure(
                    text="找不到 TXT 檔，請確認來源資料夾"))
                return
            self.after(0, lambda: self.status_lbl.configure(text="解析中，請稍候..."))
            df = parse_files(files, log_fn=lambda _: None)
            df = add_recommendation(df)
            self._df_cache = df
            counts  = df["劑型"].value_counts()
            dosages = sorted(df["劑型"].dropna().unique().tolist())
            labels  = [f"{d}（{counts.get(d, 0)}）" for d in dosages]
            self._dosage_map = dict(zip(labels, dosages))
            self.after(0, lambda: self.checklist.populate(labels, True))
            self.after(0, lambda: self.status_lbl.configure(
                text=f"共 {len(dosages)} 種劑型，{len(df):,} 筆品項"))
        except Exception as e:
            self.after(0, lambda: self.status_lbl.configure(text=f"掃描失敗：{e}"))

    # ── 執行 ────────────────────────────────────────────────
    def _run(self):
        if self.running:
            return
        sel = self.checklist.selected()
        if not sel:
            messagebox.showwarning("請選擇劑型", "請至少勾選一種劑型。")
            return
        self.running = True
        self.run_btn.configure(text="⏳ 轉換中...", bg="#4B5563")
        self.prog_bar.set_value(0, 1)
        self.log.configure(state="normal")
        self.log.delete("1.0", "end")
        self.log.configure(state="disabled")
        threading.Thread(target=self._worker, args=(sel,), daemon=True).start()

    def _worker(self, selected_labels):
        try:
            sel_dosages   = {self._dosage_map[l] for l in selected_labels
                             if l in self._dosage_map}
            out_dir       = Path(self.output_dir.get())
            is_history    = self._history_mode.get()

            files = (sorted(Path(self.input_dir.get()).glob("*.TXT")) +
                     sorted(Path(self.input_dir.get()).glob("*.txt")))

            def _progress(done, total):
                self.after(0, lambda d=done, t=total: self.prog_bar.set_value(d, t))

            # 歷史模式不使用快取（因為模式不同會產生不同資料）
            if is_history or self._df_cache is None:
                mode_txt = "（全部歷史）" if is_history else ""
                self.after(0, lambda: self._log(f"讀取 {len(files)} 個檔案{mode_txt}..."))
                df = parse_files(files,
                                 log_fn=lambda m: self.after(0, lambda msg=m: self._log(msg)),
                                 progress_fn=_progress,
                                 history_mode=is_history)
                df = add_recommendation(df)
                if not is_history:
                    self._df_cache = df  # 歷史模式結果不存快取
            else:
                df = self._df_cache

            df_out = df[df["劑型"].isin(sel_dosages)].copy()

            # 快速篩選
            filter_note = ""
            if is_history:
                filter_note = " ＋全部歷史"
            elif self._filter_future.get():
                from datetime import timedelta
                future_limit = TODAY + timedelta(days=183)
                df_out = df_out[
                    df_out["生效日"].apply(
                        lambda d: isinstance(d, date) and TODAY < d <= future_limit
                    )
                ]
                filter_note = " ＋半年內生效"
            elif self._filter_new.get():
                df_out = df_out[df_out["建議"] == "🆕 本月新增"]
                filter_note = " ＋當月更新"
            if self._filter_active.get():
                df_out = df_out[
                    df_out["停用日"].apply(
                        lambda d: not isinstance(d, date) or d > TODAY
                    )
                ]
                filter_note += " ＋有效品項"
            if self._filter_nopay.get():
                df_out = df_out[df_out["建議"] != "❌ 無健保給付"]
                filter_note += " ＋排除無給付"

            self.after(0, lambda: self._log(
                f"篩選 {len(sel_dosages)} 種劑型{filter_note} → {len(df_out):,} 筆"))

            raw_month = month_str_from_files(files)          # e.g. "202606"
            ym = f"{raw_month[:4]}-{raw_month[4:]}"          # "2026-06"
            if is_history:
                label = " 歷史資料"
            elif self._filter_future.get():
                label = " 半年內生效"
            elif self._filter_new.get():
                label = " 當月更新"
            else:
                label = ""
            nopay_suffix = " (全給付)" if self._filter_nopay.get() else ""
            ms = ym + label + nopay_suffix

            self.after(0, lambda: self._log("寫入 Excel..."))
            def _xprogress(done, total_r):
                self.after(0, lambda d=done, t=total_r: self.prog_bar.set_value(d, t))

            out_path, total, new_c = export_excel(df_out, ms, out_dir,
                                                   progress_fn=_xprogress)
            self._last_out_path = out_path

            self.after(0, lambda: self._log(
                f"\n✅ 完成！\n   {out_path.name}  ({total:,} 筆)"))
            self.after(0, lambda: self.prog_bar.stop_success())
            self.after(200, lambda: SuccessModal(self, total, new_c, out_path))

        except Exception as e:
            import traceback
            self.after(0, lambda: self._log(f"❌ {e}\n{traceback.format_exc()}"))
            self.after(0, lambda: messagebox.showerror("執行錯誤", str(e)))
            self.after(0, lambda: self.prog_bar.reset())
        finally:
            self.after(0, self._done)

    def _done(self):
        self.running = False
        self.run_btn.configure(text="▶   開始轉換", bg="#D4A017", fg="#1A1A1A")


# ════════════════════════════════════════════════════════════
#  命令列模式
# ════════════════════════════════════════════════════════════
def cli_main():
    print("=" * 50)
    print("  健保用藥品項轉換工具（命令列模式）")
    print("=" * 50)
    files = (sorted(cfg_path("input_dir").glob("*.TXT")) +
             sorted(cfg_path("input_dir").glob("*.txt")))
    if not files:
        print(f"找不到 TXT：{cfg_path('input_dir')}")
        return
    print(f"\n找到 {len(files)} 個檔案")
    df  = parse_files(files)
    df  = add_recommendation(df)
    ms  = month_str_from_files(files)
    out, total, new_c = export_excel(df, ms)
    print(f"\n完成！  {out}\n  總品項：{total}　本月新增：{new_c}")


if __name__ == "__main__":
    if "--cli" in sys.argv:
        cli_main()
    else:
        App().mainloop()
